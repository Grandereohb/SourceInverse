import argparse
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, Optional, Set, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

JING_JIN_JI_PROVINCES = {"北京市", "天津市", "河北省"}
FURNITURE_KEYWORDS = ("家具制造",)
CASTING_KEYWORDS = ("铸造",)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_problem_cell(cell) -> bool:
    fill = cell.fill
    if fill is None:
        return False

    fill_type = (fill.fill_type or fill.patternType or "").lower()
    if fill_type in {"", "none"}:
        return False

    # 默认未填充通常为 fgColor rgb=00000000，且 fill_type 为空；
    # 这里 fill_type 已过滤为空的情况，保留其余填充作为“已标色”。
    return True


def find_required_columns(headers: Iterable[object]) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for idx, raw in enumerate(headers, start=1):
        text = normalize_text(raw)
        if not text:
            continue
        if "单位名称" in text and "unit" not in header_map:
            header_map["unit"] = idx
        if "省" in text and text in {"省份", "省", "省级"} and "province" not in header_map:
            header_map["province"] = idx
        if "地市" in text and "city" not in header_map:
            header_map["city"] = idx
        if "区县" in text and "county" not in header_map:
            header_map["county"] = idx
        if "重点行业" in text and "problem_industry" not in header_map:
            header_map["problem_industry"] = idx

    required = ["unit", "city", "county", "problem_industry"]
    missing = [k for k in required if k not in header_map]
    if missing:
        raise ValueError(f"未找到必要列: {missing}，请检查表头是否包含单位名称/地市/区县/重点行业")

    return header_map


def write_xlsx(path: Path, sheet_name: str, headers: Iterable[str], rows: Iterable[Iterable[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    wb.save(path)


def contains_keywords(text: str, keywords: Iterable[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def build_bar_chart(
    ws,
    title: str,
    y_title: str,
    x_title: str,
    data_col: int,
    header_row: int,
    category_col: int,
    start_row: int,
    end_row: int,
    anchor: str,
    width: float = 12,
    height: float = 7,
) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = x_title
    chart.y_axis.delete = False
    chart.x_axis.delete = False

    values = []
    for row in range(start_row, end_row + 1):
        value = ws.cell(row, data_col).value
        if isinstance(value, (int, float)):
            values.append(value)

    max_value = max(values, default=1)
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = max(1, max_value + max(1, math.ceil(max_value * 0.1)))
    chart.y_axis.majorUnit = max(1, math.ceil(max_value / 5))

    data = Reference(ws, min_col=data_col, min_row=header_row, max_row=end_row)
    categories = Reference(ws, min_col=category_col, min_row=start_row, max_row=end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.width = width
    chart.height = height
    ws.add_chart(chart, anchor)


def write_city_county_summary_with_charts(
    path: Path,
    headers: Iterable[str],
    rows: Iterable[Iterable[object]],
    top10_cities: Iterable[Tuple[str, int]],
    top10_industries: Iterable[Tuple[str, int]],
    problem_rows: Iterable[Iterable[object]],
) -> None:
    top10_cities = list(top10_cities)
    top10_industries = list(top10_industries)
    problem_rows = [list(row) for row in problem_rows]

    jjj_rows = [row for row in problem_rows if row[0] in JING_JIN_JI_PROVINCES]
    furniture_rows = [row for row in problem_rows if contains_keywords(row[4], FURNITURE_KEYWORDS)]
    casting_rows = [row for row in problem_rows if contains_keywords(row[4], CASTING_KEYWORDS)]

    jjj_city_units: Dict[str, Set[str]] = defaultdict(set)
    furniture_city_units: Dict[str, Set[str]] = defaultdict(set)
    furniture_county_units: Dict[str, Set[str]] = defaultdict(set)
    for row in jjj_rows:
        if row[1] and row[3]:
            jjj_city_units[row[1]].add(row[3])
    for row in furniture_rows:
        if row[1] and row[3]:
            furniture_city_units[row[1]].add(row[3])
        if row[1] and row[2] and row[3]:
            furniture_county_units[f"{row[1]}-{row[2]}"].add(row[3])

    jjj_industry_counter = Counter(row[4] for row in jjj_rows if row[4])

    jjj_city_top10 = sorted(
        [(city, len(units)) for city, units in jjj_city_units.items()],
        key=lambda item: (-item[1], item[0]),
    )[:10]
    furniture_city_top10 = sorted(
        [(city, len(units)) for city, units in furniture_city_units.items()],
        key=lambda item: (-item[1], item[0]),
    )[:10]
    furniture_county_top10 = sorted(
        [(county, len(units)) for county, units in furniture_county_units.items()],
        key=lambda item: (-item[1], item[0]),
    )[:10]
    jjj_industry_top10 = jjj_industry_counter.most_common(10)

    overall_company_count = len({row[3] for row in problem_rows if row[3]})
    jjj_company_count = len({row[3] for row in jjj_rows if row[3]})
    furniture_company_count = len({row[3] for row in furniture_rows if row[3]})
    casting_company_count = len({row[3] for row in casting_rows if row[3]})

    wb = Workbook()
    ws = wb.active
    ws.title = "地市区县汇总"
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))

    focus_ws = wb.create_sheet("研究聚焦")
    focus_ws.append(["维度", "指标", "数值", "说明"])
    focus_summary_rows = [
        ["总体", "问题记录数", len(problem_rows), "标色问题行业记录总量"],
        ["总体", "问题企业数", overall_company_count, "按单位名称去重"],
        ["总体", "Top10城市", "已扩展", "按问题行业数量排序"],
        ["总体", "Top10行业", "已扩展", "按出现频次排序"],
        ["京津冀", "问题记录数", len(jjj_rows), "北京市、天津市、河北省"],
        ["京津冀", "问题企业数", jjj_company_count, "按单位名称去重"],
        ["京津冀", "涉及城市数", len({row[1] for row in jjj_rows if row[1]}), "问题记录覆盖城市"],
        ["京津冀", "涉及行业数", len({row[4] for row in jjj_rows if row[4]}), "问题行业类型"],
        ["家具制造", "问题记录数", len(furniture_rows), "问题行业含“家具制造”"],
        ["家具制造", "问题企业数", furniture_company_count, "按单位名称去重"],
        ["家具制造", "涉及城市数", len({row[1] for row in furniture_rows if row[1]}), "问题记录覆盖城市"],
        ["家具制造", "涉及区县数", len({(row[1], row[2]) for row in furniture_rows if row[1] and row[2]}), "问题记录覆盖区县"],
        ["铸造行业", "问题记录数", len(casting_rows), "问题行业含“铸造”"],
        ["铸造行业", "问题企业数", casting_company_count, "当前问题行业结果中未识别到铸造则为0"],
        ["铸造行业", "结论", "未识别到问题记录" if not casting_rows else "已识别到问题记录", "基于当前标色问题行业字段统计"],
    ]
    for row in focus_summary_rows:
        focus_ws.append(row)

    chart_ws = wb.create_sheet("图表")
    chart_ws.append(["Top10问题城市", "异常行业数量"])
    city_start_row = 2
    for city, count in top10_cities:
        chart_ws.append([city, count])

    industry_title_row = city_start_row + 12
    chart_ws.cell(row=industry_title_row, column=1, value="Top10问题行业")
    chart_ws.cell(row=industry_title_row, column=2, value="出现频次")
    industry_start_row = industry_title_row + 1
    for industry, count in top10_industries:
        chart_ws.append([industry, count])

    jjj_city_title_row = 1
    chart_ws.cell(row=jjj_city_title_row, column=10, value="京津冀问题城市Top10")
    chart_ws.cell(row=jjj_city_title_row, column=11, value="问题企业数")
    jjj_city_start_row = jjj_city_title_row + 1
    for offset, (city, count) in enumerate(jjj_city_top10):
        row_idx = jjj_city_start_row + offset
        chart_ws.cell(row=row_idx, column=10, value=city)
        chart_ws.cell(row=row_idx, column=11, value=count)

    furniture_city_title_row = industry_title_row
    chart_ws.cell(row=furniture_city_title_row, column=10, value="家具制造城市Top10")
    chart_ws.cell(row=furniture_city_title_row, column=11, value="问题企业数")
    furniture_city_start_row = furniture_city_title_row + 1
    for offset, (city, count) in enumerate(furniture_city_top10):
        row_idx = furniture_city_start_row + offset
        chart_ws.cell(row=row_idx, column=10, value=city)
        chart_ws.cell(row=row_idx, column=11, value=count)

    furniture_county_title_row = industry_title_row + 12
    chart_ws.cell(row=furniture_county_title_row, column=10, value="家具制造区县Top10")
    chart_ws.cell(row=furniture_county_title_row, column=11, value="问题企业数")
    furniture_county_start_row = furniture_county_title_row + 1
    for offset, (county, count) in enumerate(furniture_county_top10):
        row_idx = furniture_county_start_row + offset
        chart_ws.cell(row=row_idx, column=10, value=county)
        chart_ws.cell(row=row_idx, column=11, value=count)

    compare_title_row = furniture_county_title_row + 12
    chart_ws.cell(row=compare_title_row, column=10, value="研究聚焦对比")
    chart_ws.cell(row=compare_title_row, column=11, value="记录数")
    compare_start_row = compare_title_row + 1
    compare_rows = [
        ("全部问题记录", len(problem_rows)),
        ("京津冀问题记录", len(jjj_rows)),
        ("家具制造问题记录", len(furniture_rows)),
        ("铸造问题记录", len(casting_rows)),
    ]
    for offset, (label, count) in enumerate(compare_rows):
        row_idx = compare_start_row + offset
        chart_ws.cell(row=row_idx, column=10, value=label)
        chart_ws.cell(row=row_idx, column=11, value=count)

    jjj_industry_title_row = compare_title_row
    chart_ws.cell(row=jjj_industry_title_row, column=18, value="京津冀问题行业Top10")
    chart_ws.cell(row=jjj_industry_title_row, column=19, value="出现频次")
    jjj_industry_start_row = jjj_industry_title_row + 1
    for offset, (industry, count) in enumerate(jjj_industry_top10):
        row_idx = jjj_industry_start_row + offset
        chart_ws.cell(row=row_idx, column=18, value=industry)
        chart_ws.cell(row=row_idx, column=19, value=count)

    city_end_row = city_start_row + max(len(top10_cities), 1) - 1
    industry_end_row = industry_start_row + max(len(top10_industries), 1) - 1
    jjj_city_end_row = jjj_city_start_row + max(len(jjj_city_top10), 1) - 1
    furniture_city_end_row = furniture_city_start_row + max(len(furniture_city_top10), 1) - 1
    furniture_county_end_row = furniture_county_start_row + max(len(furniture_county_top10), 1) - 1
    compare_end_row = compare_start_row + len(compare_rows) - 1
    jjj_industry_end_row = jjj_industry_start_row + max(len(jjj_industry_top10), 1) - 1

    build_bar_chart(
        chart_ws, "Top10问题城市（异常行业数量）", "异常行业数量", "城市",
        2, 1, 1, city_start_row, city_end_row, "D2", width=13, height=7
    )
    build_bar_chart(
        chart_ws, "Top10问题行业（出现频次）", "出现频次", "行业",
        2, industry_title_row, 1, industry_start_row, industry_end_row, "D22", width=13, height=7
    )
    build_bar_chart(
        chart_ws, "京津冀问题城市Top10（问题企业数）", "问题企业数", "城市",
        11, jjj_city_title_row, 10, jjj_city_start_row, jjj_city_end_row, "M2", width=13, height=7
    )
    build_bar_chart(
        chart_ws, "家具制造城市Top10（问题企业数）", "问题企业数", "城市",
        11, furniture_city_title_row, 10, furniture_city_start_row, furniture_city_end_row, "M22", width=13, height=7
    )
    build_bar_chart(
        chart_ws, "家具制造区县Top10（问题企业数）", "问题企业数", "区县",
        11, furniture_county_title_row, 10, furniture_county_start_row, furniture_county_end_row, "M42", width=13, height=7
    )
    build_bar_chart(
        chart_ws, "研究聚焦对比", "记录数", "维度",
        11, compare_title_row, 10, compare_start_row, compare_end_row, "V2", width=11, height=7
    )
    if jjj_industry_top10:
        build_bar_chart(
            chart_ws, "京津冀问题行业Top10（出现频次）", "出现频次", "行业",
            19, jjj_industry_title_row, 18, jjj_industry_start_row, jjj_industry_end_row, "V22", width=11, height=7
        )

    wb.save(path)


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description="识别重点行业列中被标色的单元格，并汇总问题行业及企业数量。"
    )
    parser.add_argument(
        "excel_path",
        nargs="?",
        default="集群 - 一类集群和二类集群（终版）-20240830.xlsx",
        help="输入 Excel 文件路径",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="工作表名称（默认使用第一个工作表）",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="输出目录（默认脚本同目录）",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_path)
    if not excel_path.is_absolute():
        excel_path = script_dir / excel_path
    if not excel_path.exists():
        raise FileNotFoundError(f"找不到文件: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    cols = find_required_columns(headers)

    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = script_dir / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    problem_rows = []
    industry_frequency: Counter[str] = Counter()

    # key: (province, city, county)
    by_city_county: Dict[Tuple[str, str, str], Dict[str, Set[str]]] = defaultdict(
        lambda: {"industries": set(), "units": set()}
    )

    # key: (province, city)
    by_city: Dict[Tuple[str, str], Dict[str, Set[str]]] = defaultdict(
        lambda: {"industries": set(), "units": set(), "counties": set()}
    )

    for row_idx in range(2, ws.max_row + 1):
        industry_cell = ws.cell(row_idx, cols["problem_industry"])
        industry = normalize_text(industry_cell.value)

        if not industry:
            continue
        if not is_problem_cell(industry_cell):
            continue

        province = normalize_text(ws.cell(row_idx, cols.get("province", -1)).value) if "province" in cols else ""
        city = normalize_text(ws.cell(row_idx, cols["city"]).value)
        county = normalize_text(ws.cell(row_idx, cols["county"]).value)
        unit = normalize_text(ws.cell(row_idx, cols["unit"]).value)

        key_city_county = (province, city, county)
        by_city_county[key_city_county]["industries"].add(industry)
        if unit:
            by_city_county[key_city_county]["units"].add(unit)

        key_city = (province, city)
        by_city[key_city]["industries"].add(industry)
        by_city[key_city]["counties"].add(county)
        if unit:
            by_city[key_city]["units"].add(unit)

        problem_rows.append([province, city, county, unit, industry, row_idx])
        industry_frequency[industry] += 1

    detail_path = output_dir / "problem_records.xlsx"
    write_xlsx(
        detail_path,
        "问题记录明细",
        ["省份", "地市", "区县", "单位名称", "问题行业", "Excel行号"],
        problem_rows,
    )

    city_county_rows = []
    for (province, city, county), stats in sorted(by_city_county.items()):
        industries = sorted(stats["industries"])
        units = stats["units"]
        city_county_rows.append(
            [
                province,
                city,
                county,
                "、".join(industries),
                len(industries),
                len(units),
            ]
        )

    top10_cities = sorted(
        [(city, len(stats["industries"])) for (_, city), stats in by_city.items()],
        key=lambda x: (-x[1], x[0]),
    )[:10]
    top10_industries = industry_frequency.most_common(10)

    city_county_path = output_dir / "city_county_summary.xlsx"
    write_city_county_summary_with_charts(
        city_county_path,
        ["省份", "地市", "区县", "问题行业列表", "问题行业数量", "问题企业数量"],
        city_county_rows,
        top10_cities,
        top10_industries,
        problem_rows,
    )

    city_rows = []
    for (province, city), stats in sorted(by_city.items()):
        industries = sorted(stats["industries"])
        counties = sorted([c for c in stats["counties"] if c])
        units = stats["units"]
        city_rows.append(
            [
                province,
                city,
                "、".join(counties),
                len(counties),
                "、".join(industries),
                len(industries),
                len(units),
            ]
        )

    city_path = output_dir / "city_summary.xlsx"
    write_xlsx(
        city_path,
        "地市汇总",
        ["省份", "地市", "涉及区县列表", "涉及区县数量", "问题行业列表", "问题行业数量", "问题企业数量"],
        city_rows,
    )

    print(f"处理完成: {excel_path}")
    print(f"工作表: {ws.title}")
    print(f"识别到问题记录数: {len(problem_rows)}")
    print(f"涉及城市区县数: {len(by_city_county)}")
    print(f"涉及城市数: {len(by_city)}")
    print(f"输出文件:\n- {detail_path}\n- {city_county_path}\n- {city_path}")


if __name__ == "__main__":
    main()

